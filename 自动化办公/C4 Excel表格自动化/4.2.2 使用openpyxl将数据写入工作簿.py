"""
    File Name: 4.2.2 使用openpyxl将数据写入工作簿
    Description: openpyxl可以创建和修改工作簿，无需xlutils
    Author: 15920313768@163.com
    Date: 2022/6/27 13:23
"""

import openpyxl

# 创建工作簿对象
wb = openpyxl.Workbook()
# wb.active默认返回第一个工作表
ws = wb.active
# 第一个工作表的名称
print('ws title: ', ws.title)
# 创建一个新的worksheet
ws2 = wb.create_sheet('NewTitle', 1)
# 修改Title
ws2.title = 'MySheet'
# 添加内容
ws2.cell(row=2, column=2).value = '二两'
# 保存
wb.save('test.xlsx')

# 修改工作簿
wb = openpyxl.load_workbook('test.xlsx')
ws = wb.active
# 修改单元格
for row in ws['A2':'C3']:
    for cell in row:
        cell.value = 'new value'
wb.save('test.xlsx')
