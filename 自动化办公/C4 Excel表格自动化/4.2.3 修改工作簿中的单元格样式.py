"""
    File Name: 4.2.3 修改工作簿中的单元格样式
    Description:
    Author: 15920313768@163.com
    Date: 2022/6/27 13:29
"""

import openpyxl

# 创建工作簿对象
wb = openpyxl.Workbook()
ws = wb.active
rows = [
    ['ID', 'name', 'age'],
    ['1', 'name', 'age'],
    ['2', 'name', 'age'],
    ['3', 'name', 'age'],
]
for row in rows:
    # 添加多行
    ws.append(row)

# 通过Font修改单元格中数据的字体
from openpyxl.styles import Font, colors

# 字体设为微软雅黑，字体大小24，斜体，红色
font = Font(name='微软雅黑', size=24, italic=True, color=colors.BLUE, bold=True)
# 设置对应单元格的字体样式
ws['A1'].font = font

# 通过PatternFill修改单元格背景填充色
from openpyxl.styles import PatternFill
# 填充样式，将单元格背景色填充为绿色
fill = PatternFill(fill_type='solid', start_color=colors.BLACK)
ws['B1'].fill = fill

# 通过Border设置单元格边框样式
from openpyxl.styles import Border, Side
# 边框样式
border = Border(
    left=Side(border_style='double', color='FFBB00'),
    right=Side(border_style='double', color='FFBB00'),
    top=Side(border_style='double', color='FFBB00'),
    bottom=Side(border_style='double', color='FFBB00')
)
ws['C1'].border = border

# 通过Alignment设置单元格内容对齐方式
from openpyxl.styles import Alignment
# 单元格内容对齐
align = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws['D1'].alignment = align

# 设置单元格行高与列宽
ws.row_dimensions[3].height = 40
ws.column_dimensions['A'].width = 30

# 合并单元格
ws.merge_cells('A7:C7')
# 合并一个矩形区域的单元格
ws.merge_cells('A9:C13')
ws['A9'] = '合并单元格'
wb.save('style.xlsx')

