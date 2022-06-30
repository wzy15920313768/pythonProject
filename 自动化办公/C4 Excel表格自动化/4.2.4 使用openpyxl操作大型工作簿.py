"""
    File Name: 4.2.4 使用openpyxl操作大型工作簿
    Description: 限制内存使用
    Author: 15920313768@163.com
    Date: 2022/6/27 23:30
"""

import openpyxl
from openpyxl import load_workbook

# 在read_only模式下载入大型工作簿，将大量数据分批处理
wb = load_workbook(filename='style.xlsx', read_only=True)
# 选择工作表
ws = wb['Sheet']

for row in ws.rows:
    for cell in row:
        print(cell.value)

# write_only

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.comments import Comment
from openpyxl.styles import Font

# write_only设置为True
wb = Workbook(write_only=True)
# write_only模式下不会包含任何工作表，需要使用create_sheet方法自行创建
ws = wb.create_sheet()
# write_only模式下，单元格想要具有样式，只能使用WriteOnlyCell创建单元格
cell = WriteOnlyCell(ws, value='write_only状态写入的内容')
# 为单元格设置字体样式
cell.font = Font(name='微软雅黑', size=36)
# 插入Excel批注
cell.comment = Comment(text='这是一个批注', author='二两')
# write_only模式下只能使用append方法添加数据
ws.append([cell, 2.333, None])
# 保存
wb.save('write_only.xlsx')