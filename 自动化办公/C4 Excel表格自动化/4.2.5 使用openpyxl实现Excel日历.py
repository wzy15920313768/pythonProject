"""
    File Name: 4.2.5 使用openpyxl实现Excel日历
    Description: 通过该案例实现对Excel数据的读取、样式的修改和数据的写入
    Author: 15920313768@163.com
    Date: 2022/6/27 23:40
"""

# 导入日历库
import calendar
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font

# 指定一周的第一天， 0是星期一（默认值），6是星期天
calendar.setfirstweekday(firstweekday=6)

year = 2022
# 循环月份
for i in range(1, 13):
    # 每月中的第一行，一行表示一周
    for j in range(len(calendar.monthcalendar(year, i))):
        # 每一天
        for k in range(len(calendar.monthcalendar(year, i)[j])):
            # 具体日期
            value = calendar.monthcalendar(year, i)[j][k]

calendar.setfirstweekday(firstweekday=0)
# 创建工作簿
wb = openpyxl.Workbook()
year = 2022
for i in range(1, 13):
    # 添加工作表，每个月份对应一个工作表
    sheet = wb.create_sheet(index=0, title=str(i) + '月')
    for j in range(len(calendar.monthcalendar(year, i))):
        for k in range(len(calendar.monthcalendar(year, i)[j])):
            # 日期数据
            value = calendar.monthcalendar(year, i)[j][k]
            if value == 0:
                # 将0值变为空值，没有日期的单元格填空值
                value = ''
                sheet.cell(row=j+9, column=k+1).value = value
            else:
                # 将日期数据添加到具体的单元格中
                sheet.cell(row=j+9, column=k+1).value = value
                # 设置字体
                sheet.cell(row=j+9, column=k+1).font = Font(u'微软雅黑', size=11)
    # 单元格文字设置，右对齐，垂直居中
    align = Alignment(horizontal='right', vertical='center')
    # 单元格填充色属性设置
    fill = PatternFill('solid', fgColor='99CCCC')
    # 对单元格进行颜色填充
    for k1 in range(1, 50):
        for k2 in range(1, 50):
            sheet.cell(row=k1, column=k2).fill = fill
    # 星期日开头
    days = ['星期日', '星期一', '星期二', '星期三', '星期四', '星期五', '星期六']
    num = 0
    # 添加星期几相关信息
    for k3 in range(1, 8):
        sheet.cell(row=8, column=k3).value = days[num]
        # 设置样式
        sheet.cell(row=8, column=k3).alignment = align
        sheet.cell(row=8, column=k3).font = Font(u'微软雅黑', size=11)
        # 设置列宽12
        c_char = get_column_letter(k3)
        num += 1
    # 将日历所在单元格的行高都修改为0
    for k4 in range(8, 14):
        sheet.row_dimensions[k4].height = 30

    # 合并单元格
    sheet.merge_cells('I1:P20')
    # 添加图片
    img = Image('1.jpeg')
    # 设置图片大小
    newsize = (200, 200)
    img.width, img.height = newsize
    # 与顶部有些距离，好看一些，顶部为I1
    sheet.add_image(img, 'I2')
    # 添加年份及月份
    sheet.cell(row=3, column=1).value = f'{year}年'
    sheet.cell(row=4, column=1).value = str(i) + '月'
    # 设置年份及月份文本样式
    sheet.cell(row=3, column=1).font = Font(u'微软雅黑', size=16, bold=True, color='FF7887')
    sheet.cell(row=4, column=1).font = Font(u'微软雅黑', size=16, bold=True, color='FF7887')
    # 设置年份及月份文本对齐样式
    sheet.cell(row=3, column=1).alignment = align
    sheet.cell(row=4, column=1).alignment = align
wb.save('calendar.xlsx')


