"""
    File Name: 4.2.1 使用openpyxl读取工作簿数据
    Description:
    Author: 15920313768@163.com
    Date: 2022/6/27 12:56
"""

import openpyxl

# 打开xlsx文件
wb = openpyxl.open(r'c:\Users\extraordinary\Documents\test.xlsx')
# wb = openpyxl.load_workbook(r'c:\Users\extraordinary\Documents\test.xlsx')
# 选择第一个工作表
ws = wb.worksheets[0]
# 获取第3行第2列的值（索引从1开始）
ws.cell(row = 3, column = 2).value

# 返回最大行数
print('最大行数：', ws.max_row)
# 最小行数
print('最小行数：', ws.min_row)
# 最大列数
print('最大列数：', ws.max_column)
# 最小列数
print('最小列数：', ws.min_column)

# 遍历工作表部分区域，min_row默认从1开始
for col in ws.iter_cols(min_col=1, max_col=10, max_row=2, values_only = True):
    print(col)

# 通过ws.value也可获取工作表的值
all_values = ws.values
print(type(all_values))
for i, value in enumerate(all_values):
    print(value)
    if i == 10:
        break